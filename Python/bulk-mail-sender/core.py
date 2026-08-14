"""
Çekirdek iş mantığı: Excel okuma ve toplu gönderim döngüsü.

GUI'den bağımsızdır; komut satırından da kullanılabilir. GUI ile haberleşmek için
callback fonksiyonları (on_progress, on_log, should_stop) kullanır.

Performans notları
------------------
* Excel bir kez okunur ve (dosya yolu + değişiklik zamanı + boyut) anahtarıyla
  önbelleğe alınır. "Sütunları Oku" → "Kontrol Et" → "Gönderimi Başlat" zincirinde
  dosya tekrar ayrıştırılmaz.
* check_job'daki dosya varlık kontrolleri (os.path.isfile) iş parçacığı havuzunda
  paralel yapılır. Ağ sürücüsündeki 1300 dosya için en büyük kazanç buradadır:
  seri kontrolde her çağrı ağ gecikmesi kadar sürer.
* Gönderim kaydı (CSV) her satırda açılıp kapanmaz; iş boyunca tek bir tamponlu
  dosya kolu kullanılır.
* Bir sonraki maillerin ekleri, o an gönderim yapılırken arka planda önden okunur
  (prefetch); böylece ağdan dosya okuma süresi SMTP beklemesiyle örtüşür.
* Beklemeler time.sleep döngüsü yerine Event.wait ile yapılır: "Durdur" anında
  yanıt verir ve boş yere CPU uyandırılmaz.
"""

from __future__ import annotations

import os
import csv
import time
import threading
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass

import openpyxl

import merge
from mailer import make_mailer, MailerError


# attach_col bu değerse ek sütunu seçilmemiş demektir: mailler eksiz gider.
NO_ATTACH_COL = -1


@dataclass
class Recipient:
    row: int          # Excel'deki satır numarası (1 tabanlı, başlık dahil)
    email: str
    attachment: str
    cells: tuple = ()  # satırın TÜM hücreleri — {Sütun} yer tutucuları için


@dataclass
class SendJob:
    """Bir gönderim işinin tüm ayarlarını taşır."""
    xlsx_path: str
    sheet: str
    email_col: int          # 0 tabanlı sütun indeksi
    attach_col: int         # 0 tabanlı sütun indeksi; -1 = EK YOK (isteğe bağlıdır)
    subject: str
    body: str
    is_html: bool = False
    method: str = "smtp"    # 'smtp' | 'outlook'
    sender: str = ""
    # SMTP ayarları
    smtp_host: str = ""
    smtp_port: int = 587
    smtp_security: str = "starttls"
    smtp_user: str = ""
    smtp_password: str = ""
    # Çalışma ayarları
    delay: float = 1.0      # iki mailin BAŞLANGICI arasındaki en az süre (hız sınırı)
    start_row: int = 2      # bu satırdan itibaren (1 tabanlı; 2 = ilk veri satırı)
    limit: int = 0          # 0 = hepsi; test için ör. 3
    test_to: str = ""       # doluysa TÜM mailler bu adrese gider (güvenli test)
    results_path: str = ""  # gönderim kaydı (log) csv yolu


# ---------------------------------------------------------------------------
# Excel okuma + önbellek
# ---------------------------------------------------------------------------
_CACHE_LOCK = threading.Lock()
_HEADER_CACHE: dict = {}
_ROW_CACHE: dict = {}
_CACHE_MAX = 4          # aynı anda kaç dosyanın sonucu saklansın


def _file_stamp(path):
    """Dosyanın kimliği: (yol, değişiklik zamanı, boyut). Dosya değişince anahtar da değişir."""
    try:
        st = os.stat(path)
    except OSError:
        return None
    return (os.path.normcase(os.path.abspath(path)), st.st_mtime_ns, st.st_size)


def _cache_get(store, key):
    if key is None:
        return None
    with _CACHE_LOCK:
        return store.get(key)


def _cache_put(store, key, value):
    if key is None:
        return
    with _CACHE_LOCK:
        store[key] = value
        while len(store) > _CACHE_MAX:
            store.pop(next(iter(store)))   # en eski girdiyi at


def clear_cache():
    """Excel önbelleğini boşaltır (dosya elle değiştirildiyse gerekmez; damga kontrol edilir)."""
    with _CACHE_LOCK:
        _HEADER_CACHE.clear()
        _ROW_CACHE.clear()


def _read_rows(xlsx_path, sheet):
    """Sayfanın veri satırlarını (başlık hariç) ham demet listesi olarak döndürür.

    Sonuç önbelleğe alınır; aynı dosya/sayfa için ikinci çağrı Excel'i tekrar açmaz.
    """
    key = _file_stamp(xlsx_path)
    key = (key, sheet or "") if key else None
    hit = _cache_get(_ROW_CACHE, key)
    if hit is not None:
        return hit

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.active
        # Başlık satırı burada atılır; kalan liste 2. Excel satırından başlar.
        # (min_row=2 ile okumuyoruz: read-only modda boş satırların nasıl
        #  doldurulduğu openpyxl'in iç davranışına bağlı ve satır numaralarının
        #  Excel ile birebir aynı kalması log/CSV doğruluğu için şart.)
        rows = list(ws.iter_rows(values_only=True))[1:]
    finally:
        wb.close()

    # Çok büyük sayfaları önbellekte tutmak belleği şişirir; onları her seferinde okuruz.
    if len(rows) <= 50000:
        _cache_put(_ROW_CACHE, key, rows)
    return rows


def read_recipients(xlsx_path, sheet, email_col, attach_col=NO_ATTACH_COL):
    """Excel'i okuyup Recipient listesi döndürür (başlık satırı hariç).

    attach_col < 0 ise ek sütunu YOKTUR: her satır eksiz gönderilir. Ek yalnızca
    bir seçenektir; sadece duyuru/bilgilendirme maili atmak da geçerli bir kullanımdır.
    """
    rows = _read_rows(xlsx_path, sheet)
    recipients = []
    append = recipients.append
    ek_var = attach_col is not None and attach_col >= 0
    for idx, row in enumerate(rows, start=2):   # 2 = ilk veri satırı
        n = len(row)
        email = row[email_col] if email_col < n else None
        attach = row[attach_col] if (ek_var and attach_col < n) else None
        if email is None and attach is None:
            continue  # boş satır
        append(
            Recipient(
                row=idx,
                email=str(email).strip() if email is not None else "",
                attachment=str(attach).strip() if attach is not None else "",
                cells=row,
            )
        )
    return recipients


def read_header_names(xlsx_path, sheet):
    """Yalnızca başlıkları döndürür; okunamazsa boş liste (gönderim durmasın)."""
    try:
        _sheets, headers = get_headers(xlsx_path, sheet or None)
    except Exception:  # noqa: BLE001
        return []
    return headers


def get_headers(xlsx_path, sheet=None):
    """İlk satırı (başlıkları) ve sayfa adlarını döndürür."""
    stamp = _file_stamp(xlsx_path)
    key = (stamp, sheet or "") if stamp else None
    hit = _cache_get(_HEADER_CACHE, key)
    if hit is not None:
        return hit

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
        ws = wb[sheet] if sheet else wb.active
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = ["" if c is None else str(c) for c in row]
            break
    finally:
        wb.close()

    result = (sheet_names, headers)
    _cache_put(_HEADER_CACHE, key, result)
    return result


# ---------------------------------------------------------------------------
# Gönderim kaydı (CSV)
# ---------------------------------------------------------------------------
class _ResultWriter:
    """gonderim_sonuclari.csv'yi iş boyunca tek seferde açık tutar.

    Eski sürüm her satır için dosyayı açıp kapatıyordu; ağ/OneDrive klasöründe bu,
    mail göndermekten uzun sürebiliyordu. Ayrıca dosya Excel'de açıksa (kilitliyse)
    eski kod tüm işi çökertiyordu — artık uyarı verip kayıtsız devam eder.
    """

    def __init__(self, path, on_log=None):
        self._fh = None
        self._writer = None
        self._pending = 0
        if not path:
            return
        try:
            is_new = (not os.path.isfile(path)) or os.path.getsize(path) == 0
            self._fh = open(path, "a", encoding="utf-8-sig", newline="", buffering=1 << 16)
            self._writer = csv.writer(self._fh)
            if is_new:
                self._writer.writerow(["time", "row", "email", "status", "detail"])
        except OSError as exc:
            self._fh = None
            self._writer = None
            if on_log:
                on_log("error", f"Gönderim kaydı yazılamıyor ({exc}). Gönderim kayıtsız sürecek.")

    def write(self, row, email, status, detail, stamp=None):
        if self._writer is None:
            return
        try:
            self._writer.writerow(
                [stamp or time.strftime("%Y-%m-%d %H:%M:%S"), row, email, status, detail]
            )
            self._pending += 1
            # Program yarıda kesilirse kayıt kaybolmasın diye ara ara diske yaz.
            if self._pending >= 20:
                self._fh.flush()
                self._pending = 0
        except OSError:
            pass

    def close(self):
        if self._fh is not None:
            try:
                self._fh.flush()
                self._fh.close()
            except OSError:
                pass
        self._fh = None
        self._writer = None


# ---------------------------------------------------------------------------
# Doğrulama / ön kontrol
# ---------------------------------------------------------------------------
def validate_job(job: SendJob):
    """Göndermeden önce yaygın hataları yakalar. Sorun listesi döndürür."""
    problems = []
    if not os.path.isfile(job.xlsx_path):
        problems.append("Excel dosyası bulunamadı.")
    if not job.subject.strip():
        problems.append("Konu (başlık) boş.")
    if job.method == "smtp":
        if not job.smtp_host:
            problems.append("SMTP sunucu adresi boş.")
        if not job.sender and not job.smtp_user:
            problems.append("Gönderen adres / SMTP kullanıcı adı boş.")
    return problems


def _select_recipients(job: SendJob):
    """İşin kapsamındaki alıcıları (start_row + limit uygulanmış) döndürür."""
    recipients = read_recipients(job.xlsx_path, job.sheet, job.email_col, job.attach_col)
    if job.start_row > 2:
        recipients = [r for r in recipients if r.row >= job.start_row]
    if job.limit and job.limit > 0:
        recipients = recipients[: job.limit]
    return recipients


def _exists_map(paths, should_stop=None, on_progress=None, workers=None):
    """Verilen yolların varlığını PARALEL kontrol eder -> {yol: bool}.

    os.path.isfile bir disk/ağ çağrısıdır ve ağ sürücüsünde 10-50 ms sürebilir.
    Seri kontrolde 1300 dosya = dakikalar; havuzla saniyeler.
    """
    paths = list(paths)
    total = len(paths)
    result = {}
    if not total:
        return result
    if workers is None:
        # G/Ç beklemeli iş: çekirdek sayısından bağımsız olarak yüksek eşzamanlılık iyidir.
        workers = max(4, min(32, total))

    done = 0
    step = max(1, total // 50)      # ilerlemeyi ~50 adımda bildir (olay seli olmasın)
    with ThreadPoolExecutor(max_workers=workers) as ex:
        for path, exists in zip(paths, ex.map(os.path.isfile, paths)):
            result[path] = exists
            done += 1
            if on_progress and (done % step == 0 or done == total):
                on_progress(done, total)
            if should_stop and should_stop() and done < total:
                break
    return result


# Boş alan uyarısı köprüden JSON olarak geçtiği için sınırsız büyümesin;
# tam liste zaten kontrol_sorunlari.csv dosyasına yazılır.
_MAX_FIELD_PROBLEMS = 2000


def check_job(job: SendJob, on_progress=None, should_stop=None):
    """
    Göndermeden önce tüm listeyi tarar; hiçbir mail göndermez.

    Şunları raporlar:
      - toplam satır
      - geçersiz e-posta olan satırlar
      - ek yolu boş olan satırlar
      - ek dosyası diskte bulunamayan satırlar
      - aynı ek dosyasının birden çok satırda kullanılması (olası kopya)
      - konu/içerikte geçen ama Excel'de olmayan {alan} adları
      - {alan} kullanılan ama o satırda değeri boş olan hücreler

    Dönen sözlük 'problems' listesinde her sorun: (row, email, tip, detay)
    on_progress(done, total) : tarama ilerlemesi (isteğe bağlı)
    should_stop() -> bool    : True dönerse tarama yarıda kesilir ('stopped': True)
    """
    recipients = _select_recipients(job)
    total = len(recipients)

    bad_email = []
    empty_attach = []
    missing_attach = []
    duplicate_attach = []
    empty_field = []

    # ---- {Sütun} yer tutucuları --------------------------------------------
    mapper = merge.RowMapper(read_header_names(job.xlsx_path, job.sheet))
    subject_tpl = merge.Template(job.subject, is_html=False)
    body_tpl = merge.Template(job.body, is_html=job.is_html)
    known = mapper.keys | merge.BUILTIN_KEYS
    unknown_fields = list(dict.fromkeys(
        subject_tpl.unknown_names(known) + body_tpl.unknown_names(known)
    ))
    used_fields = list(dict.fromkeys(
        a for a in (subject_tpl.used_names() + body_tpl.used_names())
        if merge.normalize_name(a) in known
    ))
    # Varsayılanı olmayan alanlar boş kalırsa mailde boşluk görünür: bunları
    # satır satır kontrol ederiz. ({Ad|Sayın Müşterimiz} yazılmışsa sorun değil.)
    required = (subject_tpl.required_keys(known) | body_tpl.required_keys(known)) & mapper.keys
    required_names = {merge.normalize_name(n): n for n in used_fields}

    # Aynı dosya birçok satırda geçebilir; her benzersiz yolu bir kez sorgula.
    unique_paths = list(dict.fromkeys(r.attachment for r in recipients if r.attachment))
    exists = _exists_map(unique_paths, should_stop=should_stop, on_progress=on_progress)
    stopped = bool(should_stop and should_stop())

    # Ek sütunu seçilmemişse "ek boş" diye bir sorun da yoktur.
    attach_used = job.attach_col is not None and job.attach_col >= 0

    seen = {}
    norm_cache = {}
    for r in recipients:
        if not r.email or "@" not in r.email:
            bad_email.append((r.row, r.email, "gecersiz_email", r.email))
        if not r.attachment:
            # Ek İSTEĞE BAĞLIDIR: bu bir hata değil, bilgi notudur — satır
            # eksiz olarak gönderilir. Sütun hiç seçilmediyse not bile düşmeyiz.
            if attach_used:
                empty_attach.append((r.row, r.email, "ek_bos", "eksiz gönderilecek"))
        elif not exists.get(r.attachment, True):   # tarama kesildiyse "var" say
            missing_attach.append((r.row, r.email, "ek_bulunamadi", r.attachment))
        if r.attachment:
            key = norm_cache.get(r.attachment)
            if key is None:
                key = os.path.normcase(os.path.abspath(r.attachment))
                norm_cache[r.attachment] = key
            first = seen.get(key)
            if first is not None:
                duplicate_attach.append((r.row, r.email, "ek_tekrar", f"satır {first} ile aynı ek"))
            else:
                seen[key] = r.row
        if required and len(empty_field) < _MAX_FIELD_PROBLEMS:
            values = mapper.values(r.cells)
            for key in required:
                if not merge.cell_to_text(values.get(key)).strip():
                    ad = required_names.get(key, key)
                    empty_field.append((r.row, r.email, "alan_bos", f"{{{ad}}} değeri boş"))

    # 'ok' = gönderilebilir satır sayısı. Eki olmayan satır GÖNDERİLEBİLİR,
    # bu yüzden empty_attach buraya girmez (yalnızca gerçek engeller sayılır).
    ok = total - len({p[0] for p in (bad_email + missing_attach)})
    return {
        "total": total,
        "ok": ok,
        "bad_email": bad_email,
        "empty_attach": empty_attach,
        "missing_attach": missing_attach,
        "duplicate_attach": duplicate_attach,
        "empty_field": empty_field,
        "unknown_fields": unknown_fields,
        "used_fields": used_fields,
        "attach_used": attach_used,
        "stopped": stopped,
    }


# ---------------------------------------------------------------------------
# Ekleri önden okuma
# ---------------------------------------------------------------------------
class _Prefetcher(threading.Thread):
    """Gönderim sürerken bir sonraki eklerin içeriğini arka planda belleğe alır.

    Gönderim döngüsü 'advance(i)' ile nerede olduğunu bildirir; bu iş parçacığı
    en fazla 'ahead' kadar önde gider (bellek sınırlı kalsın diye).
    """

    def __init__(self, mailer, paths, stop_event, ahead=3):
        super().__init__(daemon=True, name="ek-onokuma")
        self._mailer = mailer
        self._paths = paths
        self._stop = stop_event
        self._ahead = ahead
        self._cursor = 0
        self._tick = threading.Event()
        self._cancelled = threading.Event()

    def advance(self, index):
        """Gönderim döngüsü kaçıncı satırda olduğunu bildirir."""
        self._cursor = index
        self._tick.set()

    def cancel(self):
        self._cancelled.set()
        self._tick.set()

    def _done(self):
        return self._cancelled.is_set() or self._stop.is_set()

    def run(self):
        for i, path in enumerate(self._paths):
            while i > self._cursor + self._ahead:
                if self._done():
                    return
                self._tick.wait(0.2)
                self._tick.clear()
            if self._done():
                return
            if path:
                self._mailer.prefetch(path)


# ---------------------------------------------------------------------------
# Gönderim döngüsü
# ---------------------------------------------------------------------------
def run_job(job: SendJob, on_progress=None, on_log=None, should_stop=None, stop_event=None):
    """
    Toplu gönderimi çalıştırır.

    on_progress(done, total, ok, fail) : ilerleme bildirir
    on_log(level, message)             : log satırı ('info'|'ok'|'error')
    should_stop() -> bool              : True dönerse döngü durur
    stop_event                         : threading.Event — verilirse bekleme anında kesilir

    'delay' iki mailin BAŞLANGICI arasındaki en az süredir: gönderimin kendisi
    0.8 sn sürdüyse 1 sn'lik gecikmede yalnızca 0.2 sn beklenir. Böylece sunucu
    hız sınırına uyulur ama boşa zaman harcanmaz.

    Sonuç sözlüğü döndürür: {total, sent, failed, stopped}
    """
    if stop_event is None:
        stop_event = threading.Event()
    # should_stop yalnızca stop_event'e bakıyorsa beklemeyi tek hamlede (uyanmadan)
    # yapabiliriz; dışarıdan farklı bir fonksiyon geldiyse aralıklı yoklamak gerekir.
    custom_stop = should_stop is not None and should_stop is not stop_event.is_set
    if should_stop is None:
        should_stop = stop_event.is_set

    def log(level, msg):
        if on_log:
            on_log(level, msg)

    def progress(done, total, ok, fail):
        if on_progress:
            on_progress(done, total, ok, fail)

    def stopped_now():
        return stop_event.is_set() or bool(should_stop())

    def wait_gap(seconds):
        """Mailler arası boşluk. Durdurma isteğinde anında döner. True = durduruldu."""
        if seconds <= 0:
            return stopped_now()
        if not custom_stop:
            return stop_event.wait(seconds)
        deadline = time.monotonic() + seconds
        while True:
            left = deadline - time.monotonic()
            if left <= 0:
                return False
            if stop_event.wait(min(0.15, left)) or should_stop():
                return True

    recipients = _select_recipients(job)
    total = len(recipients)
    log("info", f"Toplam işlenecek satır: {total}")
    if job.test_to:
        log("info", f"TEST MODU: tüm mailler '{job.test_to}' adresine gidecek.")

    # ---- {Sütun} yer tutucuları -------------------------------------------
    # Şablonlar BİR KEZ ayrıştırılır; her satırda yeniden çözümlenmez.
    mapper = merge.RowMapper(read_header_names(job.xlsx_path, job.sheet))
    subject_tpl = merge.Template(job.subject, is_html=False)
    body_tpl = merge.Template(job.body, is_html=job.is_html)
    known = mapper.keys | merge.BUILTIN_KEYS
    used = [a for a in (subject_tpl.used_names() + body_tpl.used_names())
            if merge.normalize_name(a) in known]
    personalize = bool(subject_tpl.fields or body_tpl.fields)
    if used:
        log("info", "Kişiselleştirme açık — kullanılan alanlar: " + ", ".join(dict.fromkeys(used)))
    for ad in dict.fromkeys(subject_tpl.unknown_names(known) + body_tpl.unknown_names(known)):
        log("error", f"UYARI: '{{{ad}}}' diye bir sütun yok; metinde olduğu gibi kalacak.")

    mailer = make_mailer(
        job.method,
        host=job.smtp_host,
        port=job.smtp_port,
        security=job.smtp_security,
        username=job.smtp_user,
        password=job.smtp_password,
        sender=job.sender,
    )
    try:
        mailer.connect()
        log("info", f"Bağlantı hazır ({job.method}).")
    except MailerError as exc:
        log("error", f"Bağlantı hatası: {exc}")
        return {"total": total, "sent": 0, "failed": 0, "stopped": True}

    results = _ResultWriter(job.results_path, on_log=on_log)

    # Ekleri önden okuma yalnızca içeriği bizim okuduğumuz yöntemde (SMTP) anlamlı.
    prefetcher = None
    if getattr(mailer, "supports_prefetch", False) and total > 1:
        prefetcher = _Prefetcher(mailer, [r.attachment for r in recipients], stop_event)
        prefetcher.start()

    ok = fail = 0
    stopped = False
    delay = max(0.0, float(job.delay or 0.0))

    try:
        for i, rcp in enumerate(recipients, start=1):
            if stopped_now():
                log("info", "Kullanıcı tarafından durduruldu.")
                stopped = True
                break
            if prefetcher is not None:
                prefetcher.advance(i)

            cycle_start = time.monotonic()
            to_addr = job.test_to or rcp.email
            stamp = time.strftime("%Y-%m-%d %H:%M:%S")

            # Doğrulamalar
            if not to_addr or "@" not in to_addr:
                fail += 1
                log("error", f"Satır {rcp.row}: geçersiz e-posta '{rcp.email}'")
                results.write(rcp.row, rcp.email, "FAIL", "gecersiz email", stamp)
                progress(i, total, ok, fail)
                continue
            if rcp.attachment and not os.path.isfile(rcp.attachment):
                fail += 1
                log("error", f"Satır {rcp.row}: ek bulunamadı -> {rcp.attachment}")
                results.write(rcp.row, rcp.email, "FAIL", "ek bulunamadi", stamp)
                progress(i, total, ok, fail)
                continue

            # Bu satıra özel konu/içerik. Değer üretimi ucuzdur ama alan yoksa
            # hiç uğraşmayız (1300 satırda gereksiz sözlük kurulmasın).
            subject, body = job.subject, job.body
            if personalize:
                values = merge.builtin_values(
                    row=rcp.row, email=rcp.email,
                    attachment=rcp.attachment, sender=job.sender,
                )
                values.update(mapper.values(rcp.cells))
                subject = subject_tpl.render(values)
                body = body_tpl.render(values)

            try:
                mailer.send(
                    to=to_addr,
                    subject=subject,
                    body=body,
                    attachment=rcp.attachment or None,
                    is_html=job.is_html,
                )
                ok += 1
                log("ok", f"Satır {rcp.row}: gönderildi -> {rcp.email}")
                results.write(rcp.row, rcp.email, "OK", "", stamp)
            except MailerError as exc:
                fail += 1
                log("error", f"Satır {rcp.row}: HATA -> {rcp.email}: {exc}")
                results.write(rcp.row, rcp.email, "FAIL", str(exc), stamp)
            except Exception as exc:  # noqa: BLE001 - tek satır tüm işi çökertmesin
                fail += 1
                log("error", f"Satır {rcp.row}: BEKLENMEYEN HATA -> {rcp.email}: {exc}")
                results.write(rcp.row, rcp.email, "FAIL", f"beklenmeyen hata: {exc}", stamp)

            progress(i, total, ok, fail)

            if delay and i < total:
                # Gönderimin kendisi kadar süre 'delay'den düşülür (hız sınırı korunur,
                # boşa beklenmez). Durdur'a basılırsa bekleme anında kesilir.
                wait_gap(delay - (time.monotonic() - cycle_start))
    finally:
        if prefetcher is not None:
            prefetcher.cancel()   # bekleyen ön-okuma iş parçacığını serbest bırak
        results.close()
        try:
            mailer.close()
        except Exception:  # noqa: BLE001
            pass

    log("info", f"Bitti. Gönderilen: {ok}, Hatalı: {fail}")
    return {"total": total, "sent": ok, "failed": fail, "stopped": stopped}
