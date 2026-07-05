"""
Çekirdek iş mantığı: Excel okuma ve toplu gönderim döngüsü.

GUI'den bağımsızdır; komut satırından da kullanılabilir. GUI ile haberleşmek için
callback fonksiyonları (on_progress, on_log, should_stop) kullanır.
"""

from __future__ import annotations

import os
import csv
import time
from dataclasses import dataclass, field

import openpyxl

from mailer import make_mailer, MailerError


@dataclass
class Recipient:
    row: int          # Excel'deki satır numarası (1 tabanlı, başlık dahil)
    email: str
    attachment: str


@dataclass
class SendJob:
    """Bir gönderim işinin tüm ayarlarını taşır."""
    xlsx_path: str
    sheet: str
    email_col: int          # 0 tabanlı sütun indeksi
    attach_col: int         # 0 tabanlı sütun indeksi
    subject: str
    body: str
    is_html: bool = False
    method: str = "smtp"    # 'smtp' | 'outlook'
    sender: str = ""
    # SMTP ayarları
    smtp_host: str = ""
    smtp_port: int = 587
    smtp_security: str = "starttls"
    smtp_user: str = ""
    smtp_password: str = ""
    # Çalışma ayarları
    delay: float = 1.0      # her mail arası saniye
    start_row: int = 2      # bu satırdan itibaren (1 tabanlı; 2 = ilk veri satırı)
    limit: int = 0          # 0 = hepsi; test için ör. 3
    test_to: str = ""       # doluysa TÜM mailler bu adrese gider (güvenli test)
    results_path: str = ""  # sonuç csv yolu (devam/resume için)
    skip_sent: bool = True  # results_path'te 'OK' olan satırları atla


def read_recipients(xlsx_path, sheet, email_col, attach_col):
    """Excel'i okuyup Recipient listesi döndürür (başlık satırı hariç)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    recipients = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue  # başlık
        email = row[email_col] if email_col < len(row) else None
        attach = row[attach_col] if attach_col < len(row) else None
        if email is None and attach is None:
            continue  # boş satır
        recipients.append(
            Recipient(
                row=idx,
                email=str(email).strip() if email is not None else "",
                attachment=str(attach).strip() if attach is not None else "",
            )
        )
    wb.close()
    return recipients


def get_headers(xlsx_path, sheet=None):
    """İlk satırı (başlıkları) ve sayfa adlarını döndürür."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    ws = wb[sheet] if sheet else wb.active
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = ["" if c is None else str(c) for c in row]
        break
    wb.close()
    return sheet_names, headers


def _load_done_rows(results_path):
    """Daha önce başarıyla gönderilmiş satır numaralarını yükler (resume)."""
    done = set()
    if results_path and os.path.isfile(results_path):
        try:
            with open(results_path, "r", encoding="utf-8-sig", newline="") as fh:
                for r in csv.DictReader(fh):
                    if (r.get("status") or "").upper() == "OK":
                        try:
                            done.add(int(r["row"]))
                        except (ValueError, KeyError, TypeError):
                            pass
        except Exception:  # noqa: BLE001
            pass
    return done


def _append_result(results_path, row, email, status, detail):
    if not results_path:
        return
    exists = os.path.isfile(results_path)
    with open(results_path, "a", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        if not exists:
            w.writerow(["time", "row", "email", "status", "detail"])
        w.writerow([time.strftime("%Y-%m-%d %H:%M:%S"), row, email, status, detail])


def validate_job(job: SendJob):
    """Göndermeden önce yaygın hataları yakalar. Sorun listesi döndürür."""
    problems = []
    if not os.path.isfile(job.xlsx_path):
        problems.append("Excel dosyası bulunamadı.")
    if not job.subject.strip():
        problems.append("Konu (başlık) boş.")
    if job.method == "smtp":
        if not job.smtp_host:
            problems.append("SMTP sunucu adresi boş.")
        if not job.sender and not job.smtp_user:
            problems.append("Gönderen adres / SMTP kullanıcı adı boş.")
    return problems


def check_job(job: SendJob):
    """
    Göndermeden önce tüm listeyi tarar; hiçbir mail göndermez.

    Şunları raporlar:
      - toplam satır
      - geçersiz e-posta olan satırlar
      - ek yolu boş olan satırlar
      - ek dosyası diskte bulunamayan satırlar
      - aynı ek dosyasının birden çok satırda kullanılması (olası kopya)

    Dönen sözlük 'problems' listesinde her sorun: (row, email, tip, detay)
    """
    recipients = read_recipients(job.xlsx_path, job.sheet, job.email_col, job.attach_col)
    recipients = [r for r in recipients if r.row >= job.start_row]
    if job.limit and job.limit > 0:
        recipients = recipients[: job.limit]

    total = len(recipients)
    bad_email = []
    empty_attach = []
    missing_attach = []
    seen = {}
    duplicate_attach = []

    for r in recipients:
        if not r.email or "@" not in r.email:
            bad_email.append((r.row, r.email, "gecersiz_email", r.email))
        if not r.attachment:
            empty_attach.append((r.row, r.email, "ek_bos", ""))
        elif not os.path.isfile(r.attachment):
            missing_attach.append((r.row, r.email, "ek_bulunamadi", r.attachment))
        if r.attachment:
            key = os.path.normcase(os.path.abspath(r.attachment))
            if key in seen:
                duplicate_attach.append((r.row, r.email, "ek_tekrar", f"satır {seen[key]} ile aynı ek"))
            else:
                seen[key] = r.row

    ok = total - len({p[0] for p in (bad_email + empty_attach + missing_attach)})
    return {
        "total": total,
        "ok": ok,
        "bad_email": bad_email,
        "empty_attach": empty_attach,
        "missing_attach": missing_attach,
        "duplicate_attach": duplicate_attach,
    }


def run_job(job: SendJob, on_progress=None, on_log=None, should_stop=None):
    """
    Toplu gönderimi çalıştırır.

    on_progress(done, total, ok, fail) : ilerleme bildirir
    on_log(level, message)             : log satırı ('info'|'ok'|'error')
    should_stop() -> bool              : True dönerse döngü durur

    Sonuç sözlüğü döndürür: {total, sent, failed, skipped, stopped}
    """
    def log(level, msg):
        if on_log:
            on_log(level, msg)

    def progress(done, total, ok, fail):
        if on_progress:
            on_progress(done, total, ok, fail)

    recipients = read_recipients(job.xlsx_path, job.sheet, job.email_col, job.attach_col)

    # start_row filtresi
    recipients = [r for r in recipients if r.row >= job.start_row]

    done_rows = _load_done_rows(job.results_path) if job.skip_sent else set()

    if job.limit and job.limit > 0:
        recipients = recipients[: job.limit]

    total = len(recipients)
    log("info", f"Toplam işlenecek satır: {total}")
    if job.test_to:
        log("info", f"TEST MODU: tüm mailler '{job.test_to}' adresine gidecek.")

    mailer = make_mailer(
        job.method,
        host=job.smtp_host,
        port=job.smtp_port,
        security=job.smtp_security,
        username=job.smtp_user,
        password=job.smtp_password,
        sender=job.sender,
    )
    try:
        mailer.connect()
        log("info", f"Bağlantı hazır ({job.method}).")
    except MailerError as exc:
        log("error", f"Bağlantı hatası: {exc}")
        return {"total": total, "sent": 0, "failed": 0, "skipped": 0, "stopped": True}

    ok = fail = skipped = 0
    stopped = False

    for i, rcp in enumerate(recipients, start=1):
        if should_stop and should_stop():
            log("info", "Kullanıcı tarafından durduruldu.")
            stopped = True
            break

        if rcp.row in done_rows:
            skipped += 1
            progress(i, total, ok, fail)
            continue

        to_addr = job.test_to or rcp.email

        # Doğrulamalar
        if not to_addr or "@" not in to_addr:
            fail += 1
            log("error", f"Satır {rcp.row}: geçersiz e-posta '{rcp.email}'")
            _append_result(job.results_path, rcp.row, rcp.email, "FAIL", "gecersiz email")
            progress(i, total, ok, fail)
            continue
        if rcp.attachment and not os.path.isfile(rcp.attachment):
            fail += 1
            log("error", f"Satır {rcp.row}: ek bulunamadı -> {rcp.attachment}")
            _append_result(job.results_path, rcp.row, rcp.email, "FAIL", "ek bulunamadi")
            progress(i, total, ok, fail)
            continue

        try:
            mailer.send(
                to=to_addr,
                subject=job.subject,
                body=job.body,
                attachment=rcp.attachment or None,
                is_html=job.is_html,
            )
            ok += 1
            log("ok", f"Satır {rcp.row}: gönderildi -> {rcp.email}")
            _append_result(job.results_path, rcp.row, rcp.email, "OK", "")
        except MailerError as exc:
            fail += 1
            log("error", f"Satır {rcp.row}: HATA -> {rcp.email}: {exc}")
            _append_result(job.results_path, rcp.row, rcp.email, "FAIL", str(exc))

        progress(i, total, ok, fail)

        if job.delay and i < total:
            # Durdurmaya duyarlı bekleme
            waited = 0.0
            step = 0.1
            while waited < job.delay:
                if should_stop and should_stop():
                    break
                time.sleep(step)
                waited += step

    mailer.close()
    log("info", f"Bitti. Gönderilen: {ok}, Hatalı: {fail}, Atlanan: {skipped}")
    return {"total": total, "sent": ok, "failed": fail, "skipped": skipped, "stopped": stopped}
